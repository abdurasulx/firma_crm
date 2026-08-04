"""
Excel export views — savdolar, nasiya, mahsulotlar, xodimlar
"""
import io
import datetime as dt
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Sum

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Stil yordamchilari ────────────────────────────────────────────────────────

_GREEN  = "FF10B981"
_DARK   = "FF0F172A"
_LIGHT  = "FFF8FAFC"
_YELLOW = "FFFFF3CD"
_RED    = "FFFEE2E2"

def _header_style(ws, row, cols):
    """Row'ni yashil header qilib formatlash"""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = Font(bold=True, color="FFFFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor=_GREEN)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="FFCCCCCC")
        cell.border    = Border(left=thin, right=thin, bottom=thin, top=thin)

def _title_row(ws, text, cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font      = Font(bold=True, size=14, color=_DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill      = PatternFill("solid", fgColor="FFE2F5ED")
    ws.row_dimensions[1].height = 30

def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

def _response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ── 1. SAVDOLAR EXPORT ────────────────────────────────────────────────────────

@login_required(login_url='login')
def export_savdolar(request):
    from .models import Savdo
    if request.user.type not in ('ega', 'savdogar'):
        from django.shortcuts import redirect
        return redirect('main')

    company = request.company
    now     = timezone.localtime()

    date_from_str = request.GET.get('from')
    date_to_str   = request.GET.get('to')

    try:
        date_from = dt.datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else (now.date() - dt.timedelta(days=30))
        date_to   = dt.datetime.strptime(date_to_str,   '%Y-%m-%d').date() if date_to_str   else now.date()
    except ValueError:
        date_from = now.date() - dt.timedelta(days=30)
        date_to   = now.date()

    start = timezone.make_aware(dt.datetime.combine(date_from, dt.time.min))
    end   = timezone.make_aware(dt.datetime.combine(date_to,   dt.time.max))

    qs = Savdo.objects.filter(
        company=company, vaqt_sana__range=(start, end)
    ).select_related('haridor_dukon', 'yetkazib_beruvchi__user', 'savdogar').order_by('-vaqt_sana')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Savdolar"

    HEADERS = ["#", "Sana", "Mijoz", "Yetkazuvchi / Savdogar", "To'lov turi", "Summa (so'm)", "Holat"]
    _title_row(ws, f"{company.name} — Savdolar ({date_from.strftime('%d.%m.%Y')} – {date_to.strftime('%d.%m.%Y')})", len(HEADERS))
    for col, h in enumerate(HEADERS, 1):
        ws.cell(row=2, column=col, value=h)
    _header_style(ws, 2, len(HEADERS))

    for i, s in enumerate(qs, 1):
        customer = s.haridor_dukon.nomi if s.haridor_dukon else s.oluvchining_ismi
        agent    = ""
        if s.savdogar:
            agent = s.savdogar.tuliq_ismi or s.savdogar.username
        elif s.yetkazib_beruvchi:
            agent = s.yetkazib_beruvchi.user.tuliq_ismi or s.yetkazib_beruvchi.user.username
        status   = "To'landi" if s.tulandi else ("Nasiya" if s.st == "nasiya" else "—")
        row      = i + 2
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=timezone.localtime(s.vaqt_sana).strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=3, value=customer)
        ws.cell(row=row, column=4, value=agent)
        ws.cell(row=row, column=5, value=s.get_st_display())
        ws.cell(row=row, column=6, value=float(s.summa or 0))
        ws.cell(row=row, column=7, value=status)
        if s.st == 'nasiya' and not s.tulandi:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFFFF3CD")
        ws.row_dimensions[row].height = 18

    # Jami
    total_row = len(list(qs)) + 3
    ws.cell(row=total_row, column=5, value="JAMI:").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=float(qs.aggregate(t=Sum('summa'))['t'] or 0)).font = Font(bold=True)

    _auto_width(ws)
    filename = f"savdolar_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.xlsx"
    return _response(wb, filename)


# ── 2. NASIYA EXPORT ──────────────────────────────────────────────────────────

@login_required(login_url='login')
def export_nasiya(request):
    from .models import Savdo, NasiyaTolov
    if request.user.type not in ('ega', 'savdogar'):
        from django.shortcuts import redirect
        return redirect('main')

    company = request.company
    nasiya_qs = Savdo.objects.filter(
        company=company, st='nasiya', tulandi=False
    ).select_related('haridor_dukon', 'yetkazib_beruvchi__user').order_by('credit_due_date')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nasiya"

    HEADERS = ["#", "Mijoz", "Yetkazuvchi", "Savdo summasi", "To'langan", "Qoldiq", "Muddat", "Holat"]
    _title_row(ws, f"{company.name} — Ochiq Nasiyalar", len(HEADERS))
    for col, h in enumerate(HEADERS, 1):
        ws.cell(row=2, column=col, value=h)
    _header_style(ws, 2, len(HEADERS))

    today = timezone.localtime().date()
    total_debt = 0.0

    for i, s in enumerate(nasiya_qs, 1):
        customer = s.haridor_dukon.nomi if s.haridor_dukon else s.oluvchining_ismi
        agent = ""
        if s.yetkazib_beruvchi:
            agent = s.yetkazib_beruvchi.user.tuliq_ismi or s.yetkazib_beruvchi.user.username
        tolangan = NasiyaTolov.objects.filter(savdo=s).aggregate(t=Sum('tolov_summasi'))['t'] or 0
        qoldiq   = float(s.summa or 0) - float(tolangan)
        if qoldiq <= 0:
            continue

        due_date = s.credit_due_date
        if due_date:
            if due_date < today:
                holat = f"⚠ {(today - due_date).days} kun kechikdi"
            elif (due_date - today).days <= 3:
                holat = f"⏰ {(due_date - today).days} kun qoldi"
            else:
                holat = f"✓ {(due_date - today).days} kun qoldi"
        else:
            holat = "Muddat belgilanmagan"

        total_debt += qoldiq
        row = i + 2
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=customer)
        ws.cell(row=row, column=3, value=agent)
        ws.cell(row=row, column=4, value=float(s.summa or 0))
        ws.cell(row=row, column=5, value=float(tolangan))
        ws.cell(row=row, column=6, value=qoldiq)
        ws.cell(row=row, column=7, value=due_date.strftime('%d.%m.%Y') if due_date else "—")
        ws.cell(row=row, column=8, value=holat)

        if due_date and due_date < today:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFFEE2E2")
        elif due_date and (due_date - today).days <= 3:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFFFF3CD")
        ws.row_dimensions[row].height = 18

    total_row = nasiya_qs.count() + 3
    ws.cell(row=total_row, column=5, value="JAMI QOLDIQ:").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=round(total_debt, 2)).font = Font(bold=True, color="FFDC2626")

    _auto_width(ws)
    today_str = today.strftime('%Y%m%d')
    return _response(wb, f"nasiya_{today_str}.xlsx")


# ── 3. MAHSULOTLAR EXPORT ─────────────────────────────────────────────────────

@login_required(login_url='login')
def export_mahsulotlar(request):
    from .models import Mahsulot
    if request.user.type != 'ega':
        from django.shortcuts import redirect
        return redirect('main')

    company = request.company
    qs = Mahsulot.objects.filter(company=company).select_related('turi').order_by('nomi')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mahsulotlar"

    HEADERS = ["#", "Nomi", "Turi", "Narxi (so'm)", "Zaxira miqdori", "Min. miqdor", "Holat"]
    _title_row(ws, f"{company.name} — Mahsulotlar", len(HEADERS))
    for col, h in enumerate(HEADERS, 1):
        ws.cell(row=2, column=col, value=h)
    _header_style(ws, 2, len(HEADERS))

    for i, p in enumerate(qs, 1):
        holat = "⚠ Kam" if p.miqdori < p.min_miqdori else "✓ Normal"
        row = i + 2
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=p.nomi)
        ws.cell(row=row, column=3, value=p.turi.nomi)
        ws.cell(row=row, column=4, value=float(p.narxi))
        ws.cell(row=row, column=5, value=float(p.miqdori))
        ws.cell(row=row, column=6, value=float(p.min_miqdori))
        ws.cell(row=row, column=7, value=holat)
        if p.miqdori < p.min_miqdori:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFFEE2E2")
        ws.row_dimensions[row].height = 18

    _auto_width(ws)
    return _response(wb, f"mahsulotlar_{timezone.localtime().strftime('%Y%m%d')}.xlsx")


# ── 4. XODIMLAR EXPORT ────────────────────────────────────────────────────────

@login_required(login_url='login')
def export_xodimlar(request):
    from .models import User
    if request.user.type != 'ega':
        from django.shortcuts import redirect
        return redirect('main')

    company = request.company
    qs = User.objects.filter(company=company).exclude(type='ega').order_by('type', 'tuliq_ismi')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Xodimlar"

    HEADERS = ["#", "To'liq ismi", "Username", "Lavozimi", "Tel raqami", "Faollik", "Qo'shilgan sana"]
    _title_row(ws, f"{company.name} — Xodimlar", len(HEADERS))
    for col, h in enumerate(HEADERS, 1):
        ws.cell(row=2, column=col, value=h)
    _header_style(ws, 2, len(HEADERS))

    TYPE_MAP = {
        'yetkazib_beruvchi': 'Yetkazuvchi',
        'pazanda': 'Ishlab chiqaruvchi',
        'omborchi': 'Omborchi',
        'savdogar': 'Savdogar',
        'ishlab_chiqaruvchi': 'Ishlab chiqaruvchi',
    }

    for i, u in enumerate(qs, 1):
        row = i + 2
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=u.tuliq_ismi or u.username)
        ws.cell(row=row, column=3, value=u.username)
        ws.cell(row=row, column=4, value=TYPE_MAP.get(u.type, u.type))
        ws.cell(row=row, column=5, value=u.tel_raqami or "—")
        ws.cell(row=row, column=6, value="Faol" if u.is_active else "Nofaol")
        ws.cell(row=row, column=7, value=u.date_joined.strftime('%d.%m.%Y') if u.date_joined else "—")
        if not u.is_active:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFF1F5F9")
        ws.row_dimensions[row].height = 18

    _auto_width(ws)
    return _response(wb, f"xodimlar_{timezone.localtime().strftime('%Y%m%d')}.xlsx")
