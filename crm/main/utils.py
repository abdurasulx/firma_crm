import math
import os
import pandas as pd
import io
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


def haversine_metres(lat1, lng1, lat2, lng2):
    """Ikki GPS nuqta orasidagi masofa (metrda) — Yer sferasi bo'yicha
    (haversine formulasi). Qisqa masofalarda (bir necha yuz metrgacha)
    yetarli aniqlikda."""
    R = 6371000.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    d_phi = math.radians(lat2 - lat1)
    d_lambda = math.radians(lng2 - lng1)
    a = math.sin(d_phi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(d_lambda / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def validate_uploaded_file(f, allowed_ext=('.pdf', '.jpg', '.jpeg', '.png'), max_mb=10):
    """Oddiy `FileField`lar (masalan shartnoma PDF'i) uchun — `ImageField`
    kabi Pillow orqali avtomatik tekshirilmaydi, shuning uchun kengaytma va
    hajmni qo'lda tekshirish kerak. Fayl noto'g'ri bo'lsa `ValidationError`
    ko'taradi, chaqiruvchi buni forma xatosi sifatida ko'rsatishi kerak."""
    ext = os.path.splitext(f.name)[1].lower()
    if ext not in allowed_ext:
        raise ValidationError(f"Ruxsat etilmagan fayl turi: {ext}")
    if f.size > max_mb * 1024 * 1024:
        raise ValidationError(f"Fayl hajmi {max_mb}MB dan katta")

def format_product_string(smm_str, company=None):
    """
    Transforms "Product1 Qty1 Price1, Product2 Qty2 Price2" 
    into "Product1 Qty1 Unit1 (Price1), Product2 Qty2 Unit2 (Price2)"
    """
    from .models import Mahsulot
    if not smm_str:
        return ""
    
    formatted_items = []
    # Split by comma first
    items = [i.strip() for i in smm_str.split(',') if i.strip()]
    
    for item in items:
        # Expected format: "Product Name Quantity Price"
        # We use rsplit to handle product names with spaces
        parts = item.rsplit(' ', 2)
        if len(parts) == 3:
            name, qty, price = parts
            name = name.strip()
            # Try to find the product to get its unit (turi)
            p = Mahsulot.objects.filter(nomi=name, company=company).first()
            unit = p.turi if (p and p.turi) else ""
            formatted_items.append(f"{name} {qty} {unit} ({price})".replace("  ", " ").strip())
        elif len(parts) == 2:
            name, qty = parts
            name = name.strip()
            p = Mahsulot.objects.filter(nomi=name, company=company).first()
            unit = p.turi if (p and p.turi) else ""
            formatted_items.append(f"{name} {qty} {unit}".replace("  ", " ").strip())
        else:
            formatted_items.append(item)
            
    return ", ".join(formatted_items)

def export_to_excel(df, filename, header_info=None):
    """
    Exports a DataFrame to Excel with custom styling:
    - Headers for user info and date range
    - Table borders
    - Auto-filters
    - Wrapped text for Mahsulotlar
    """
    output = io.BytesIO()
    
    # Create a writer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # We start writing the data from row 4 (1-indexed for openpyxl, but 0-indexed for pandas is complicated)
        # So we'll write to row 3 (0-indexed) which is row 4 in Excel
        df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=3)
        
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # 1. Header Information
        if header_info:
            # Row 1: Title or User Name
            worksheet.merge_cells('A1:F1')
            cell_a1 = worksheet['A1']
            cell_a1.value = header_info.get('title', 'Hisobot')
            cell_a1.font = Font(size=14, bold=True)
            cell_a1.alignment = Alignment(horizontal='center')
            
            # Row 2: Date Range
            worksheet.merge_cells('A2:F2')
            cell_a2 = worksheet['A2']
            cell_a2.value = f"Davr: {header_info.get('date_range', '-')}"
            cell_a2.font = Font(size=11, italic=True)
            cell_a2.alignment = Alignment(horizontal='center')

        # 2. Table Formatting (Starts at row 4)
        thin_side = Side(border_style="thin", color="000000")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center')
        
        # Iterate through rows and columns to apply borders and alignment
        # Row 4 is the header row
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        for row in range(4, max_row + 1):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border
                if row == 4:
                    cell.font = header_font
                    cell.alignment = center_alignment
                
                # Special handling for 'Mahsulotlar' or long text
                if col == 3: # Assuming Mahsulotlar is column 3
                    cell.alignment = Alignment(wrap_text=True)

        # 3. Auto-filter
        worksheet.auto_filter.ref = f"A4:{get_column_letter(max_col)}{max_row}"
        
        # 4. Column Widths (Adjust as needed)
        worksheet.column_dimensions['A'].width = 12 # Sana
        worksheet.column_dimensions['B'].width = 10 # Vaqt
        worksheet.column_dimensions['C'].width = 20 # Haridor / Ism
        worksheet.column_dimensions['D'].width = 40 # Mahsulotlar / Mahsulot
        worksheet.column_dimensions['E'].width = 15 # Summa / Miqdor
        worksheet.column_dimensions['F'].width = 15 # To'lov / Izoh
        if max_col > 6:
            worksheet.column_dimensions['G'].width = 10 # Holati (for savdolar)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
